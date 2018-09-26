import openpyxl
import requests
import os
import datetime as dt
import logging
import sys

if sys.version_info[0] < 3:
    raise ImportError('Python < 3 is not supported.')

if not os.path.isfile(os.path.join('.', 'add_documents.py')):
    raise RuntimeError('Must run this script from the directory it is in.')

log_filepath = '_logs/{}.log'.format(dt.datetime.utcnow().strftime('%Y-%m-%d_%H-%M-%S'))
log_format = '%(asctime)s - %(levelname)s - %(message)s'
logging.basicConfig(  # filename=log_filepath,
                    format=log_format, level=logging.INFO)
logging.getLogger("requests").setLevel(logging.WARNING)
logging.getLogger("urllib3").setLevel(logging.WARNING)

settings = {
    'namenindex': {
        'keys': ['datum', 'naam', 'inhoud', 'bron', 'getuigen', 'bijzonderheden'],
        'url': 'http://localhost:9200/namenindex/doc/{}'
    },
    'maatboek_heemskerk': {
        'keys': ['gebied', 'sector', 'nummer', 'oppervlakte', 'eigenaar', 'huurder',
                 'bedrag', 'jaar', 'bron', 'opmerkingen'],
        'url': 'http://localhost:9200/maatboek_heemskerk/doc/{}'
    }
}
session = requests.Session()

for filename in sorted(os.listdir('.')):
    if not filename.endswith('.xlsx'):
        continue
    if 'maatboek' in filename.lower() and 'heemskerk' in filename.lower():
        keys = settings['maatboek_heemskerk']['keys']
        url = settings['maatboek_heemskerk']['url']
    else:
        keys = settings['namenindex']['keys']
        url = settings['namenindex']['url']
    logging.info('Working on {}.'.format(filename))
    wb = openpyxl.load_workbook(filename, read_only=True)
    sheet = wb.active
    rowmax = sheet.max_row
    count_no_id = 0
    if rowmax < 10:
        logging.warning('The max row is suspiciously low: {}.'.format(rowmax))
        continue  # to next file
    for i in range(2, rowmax):
        # if this row is not available go to the next workbook
        try:
            if sheet.cell(row=i, column=1).value is None:
                count_no_id += 1
                if count_no_id > 10:
                    logging.info('Id is None at too many rows, lastly at i = {}, go to next file.'
                                 .format(i))
                    break
        except IndexError:
            logging.warning('IndexError at i = {}, go to next file.'.format(i))
            break
        count_no_id = 0
        if i % 250 == 0:
            logging.info('Working on row {} of {}.'.format(i, filename))
        doc_id = sheet.cell(row=i, column=1).value
        # get data from cells that have a value
        doc = {}
        for j, key in enumerate(keys):
            cell_value = sheet.cell(row=i, column=j + 2).value
            if cell_value is not None:
                doc[key] = str(cell_value).lstrip().rstrip()
        # skip row if there is no data except an id
        if len(doc.keys()) == 0:
            continue
        # split name in normal and keyword version (only for namenindex, so check on unique key)
        if 'naam' in doc:
            # fix this one: Albrecht (St), van
            if len(doc['naam'].split(',')) >= 2:
                doc['naam_keyword'] = doc['naam'].split(',')[0]
            elif len(doc['naam'].split('~')) >= 2:
                doc['naam_keyword'] = doc['naam'].split('~')[0]
            elif len(doc['naam'].split(' ')) >= 2:
                doc['naam_keyword'] = doc['naam'].split(' ')[0]
            else:
                doc['naam_keyword'] = doc['naam']
        # start sending data        
        resp = session.put(url.format(doc_id), json=doc)
        try:
            resp.raise_for_status()
        except requests.exceptions.HTTPError:
            fill = 'HTTP error: code {} at i={} in {}.'
            logging.warning(fill.format(resp.status_code, i, filename))
            break

logging.info("Finished!")
