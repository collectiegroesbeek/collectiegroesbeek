""""
This seems to be a temporary script to develop adding Maatboek Heemskerk.
"""

import openpyxl
import requests
import os
import datetime as dt
import logging
import sys

if sys.version_info[0] < 3:
    raise ImportError('Python < 3 is unsupported.')

log_filepath = 'logs/{}.log'.format(dt.datetime.utcnow().strftime('%Y-%m-%d_%H-%M-%S'))
log_format = '%(asctime)s - %(levelname)s - %(message)s'
logging.basicConfig(filename=log_filepath, format=log_format, level=logging.INFO)
logging.getLogger("requests").setLevel(logging.WARNING)
logging.getLogger("urllib3").setLevel(logging.WARNING)

filename = 'Heemskerk maatboek.xlsx'

session = requests.Session()

logging.info('Working on {}.'.format(filename))
wb = openpyxl.load_workbook(filename, read_only=True)
sheet = wb.active
rowmax = sheet.max_row
if rowmax < 10:
    logging.warning('The max row is suspiciously low: {}.'.format(rowmax))

for i in range(2, rowmax):
    if i > 30:
        break
    # if this row is not available go to the next workbook
    try:
        if sheet.cell(row=i, column=1).value is None:
            logging.info('Id is None at i = {}, stopping.'.format(i))
            break
    except IndexError:
        logging.warning('IndexError at i = {}, stopping.'.format(i))
        break
    if i % 250 == 0:
        logging.info('Working on row {} of {}.'.format(i, filename))

    doc_id = sheet.cell(row=i, column=1).value
    # get data from cells that have a value
    doc = {}
    for j, key in enumerate(keys):
        cell_value = sheet.cell(row=i, column=j+1).value
        if cell_value is not None:
            doc[key] = str(cell_value).lstrip().rstrip()
    # skip row if there is very little data
    if len(doc.keys()) <= 1:
        continue
    # start sending data
    resp = session.put(url.format(doc_id), json=doc)
    try:
        resp.raise_for_status()
    except requests.exceptions.HTTPError:
        fill = 'HTTP error: code {} at i={} in {}.'
        logging.warning(fill.format(resp.status_code, i, filename))
        break

logging.info("Finished!")
